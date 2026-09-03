import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

# 1. عنوان الصفحة
st.set_page_config(page_title="مُحول الملفات - الشركة", page_icon="📊", layout="centered")

st.title("📊 أداة تحويل ملفات HTML إلى Excel")
st.write("مرحباً بك! قم برفع ملف الـ HTML الخاص بالتقارير، وسيقوم النظام بتحويله تلقائياً إلى ملف Excel مُنسق وجاهز.")

def convert_html_to_excel(html_bytes):
    tables = pd.read_html(html_bytes)
    
    wb = Workbook()
    wb.remove(wb.active) # حذف الورقة الافتراضية

    # ألوان التنسيق الخاصة بالجدول
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for i, df in enumerate(tables):
        ws = wb.create_sheet(title=f"تقرير_{i+1}")
        ws.views.sheetView[0].showGridLines = True

        ws.append(list(df.columns))

        for row in df.itertuples(index=False):
            ws.append(list(row))

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            for cell in row:
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if r_idx % 2 == 0:
                    cell.fill = zebra_fill

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len * 2, 14)

        ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# 2. زر رفع الملف
uploaded_file = st.file_uploader("اختر ملف الـ HTML من جهازك", type=["html", "htm"])

if uploaded_file is not None:
    try:
        with st.spinner('جاري معالجة الملف وتنسيقه...'):
            excel_data = convert_html_to_excel(uploaded_file)
            
        st.success("🎉 تم تحويل الملف بنجاح!")
        
        # 3. زر التحميل
        st.download_button(
            label="📥 تحميل ملف Excel الناتج",
            data=excel_data,
            file_name="converted_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        st.error(f"حدث خطأ أثناء معالجة الملف، تأكد أن الملف يحتوي على جداول صحيحة.")